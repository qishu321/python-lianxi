import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# MySQL数据库连接信息
host = '192.168.2.81'
user = 'root'
password = '123456'
database = 'cmdb-ops-flow'

# 建立MySQL数据库连接
connection = pymysql.connect(host=host, user=user, password=password, database=database)
cursor = connection.cursor()

# 获取所有表名
cursor.execute("SHOW TABLES")
tables = cursor.fetchall()

# 创建Excel文件
wb = Workbook()

# 创建一个工作表用于统计表名
summary_sheet = wb.active
summary_sheet.title = 'Summary'
summary_sheet.append(['Table Name', 'Table Description', 'Link'])

# 循环遍历表并添加到Excel文件
for table in tables:
    table_name = table[0]

    # 获取表的描述信息，如果没有描述信息，使用默认值
    cursor.execute(
        f"SELECT table_comment FROM information_schema.tables WHERE table_schema='{database}' AND table_name='{table_name}'")
    result = cursor.fetchone()
    table_description = result[0] if result else 'No description available'

    # 创建一个新的工作表，并将表名作为标题
    sheet = wb.create_sheet(title=table_name)
    sheet.append(['Field Name', 'Field Type', 'Field Description', 'Nullable', 'Primary Key', 'Auto Increment'])

    # 获取表的字段信息
    cursor.execute(f"SHOW FULL COLUMNS FROM {table_name}")
    columns = cursor.fetchall()

    for column in columns:
        field_name = column[0]
        field_type = column[1]
        field_description = column[8]  # 字段注释
        nullable = 'YES' if column[3] == 'YES' else 'NO'
        primary_key = 'YES' if 'PRI' in column[4] else 'NO'
        auto_increment = 'YES' if column[5] and 'auto_increment' in column[5] else 'NO'
        sheet.append([field_name, field_type, field_description, nullable, primary_key, auto_increment])

    # 设置工作表的标题行为粗体
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # 在"Summary"工作表中创建一个超链接单元格，将其值设置为表名，超链接设置为跳转到对应的表工作表
    # summary_link_cell = f"A{len(summary_sheet['A']) + 1}"
    # summary_sheet[summary_link_cell] = table_name
    # summary_sheet[summary_link_cell].hyperlink = f"'{table_name}'!A1"

    # 在"Summary"工作表中添加表名和表描述
    # summary_sheet.append([table_name, table_description, summary_sheet[summary_link_cell]])

    # 填充Summary页中的信息
    table_link = f"=HYPERLINK(\"#{table_name}!A1\",\"跳转\")"
    summary_sheet.append([table_name, table_description, table_link])

# 关闭数据库连接
cursor.close()
connection.close()

# 保存Excel文件
wb.save('mysql_tables_info2.xlsx')
